import streamlit as st

# Page configuration - MUST be the first Streamlit command
st.set_page_config(
    page_title="Financial Valuation Engine",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

import pandas as pd
import numpy as np
import json
from datetime import datetime
import io

# Visualization - use matplotlib as primary
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for deployment

# Optional imports for enhanced visualization
try:
    import plotly.graph_objects as go
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    # No warning needed since matplotlib is primary

# Optional Excel export functionality
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    st.warning("Excel export not available. CSV download only.")

from params import ValuationParams
from valuation import calc_dcf_series, calc_apv
from montecarlo import run_monte_carlo
from multiples import run_multiples_analysis
from scenario import run_scenarios, detect_circular_references
from sensitivity import run_sensitivity_analysis

# Custom CSS for professional appearance
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .author {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
        font-style: italic;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
    }
    .success-message {
        background-color: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 0.5rem;
        border: 1px solid #c3e6cb;
    }
    .error-message {
        background-color: #f8d7da;
        color: #721c24;
        padding: 1rem;
        border-radius: 0.5rem;
        border: 1px solid #f5c6cb;
    }
</style>
""", unsafe_allow_html=True)



def main():
    """Main application function"""
    
    # Header
    st.markdown('<h1 class="main-header">Financial Valuation Engine</h1>', unsafe_allow_html=True)
    st.markdown('<p class="author">By Pranay Upreti</p>', unsafe_allow_html=True)
    
    # Initialize session state
    if 'valuation_results' not in st.session_state:
        st.session_state.valuation_results = None
    if 'analysis_type' not in st.session_state:
        st.session_state.analysis_type = []
    
    # Sidebar for navigation
    with st.sidebar:
        st.header("Analysis Configuration")
        
        # Analysis selection
        st.subheader("Select Analyses")
        run_wacc = st.checkbox("WACC DCF", value=True, help="Standard DCF using WACC")
        run_apv = st.checkbox("APV DCF", value=False, help="Adjusted Present Value method")
        do_mc = st.checkbox("Monte Carlo", value=False, help="Uncertainty analysis")
        do_mult = st.checkbox("Comparable Multiples", value=False, help="Peer company analysis")
        do_scen = st.checkbox("Scenario Analysis", value=False, help="What-if scenarios")
        do_sens = st.checkbox("Sensitivity Analysis", value=False, help="Parameter sensitivity")
        
        # Store selected analyses
        selected_analyses = []
        if run_wacc: selected_analyses.append("WACC DCF")
        if run_apv: selected_analyses.append("APV DCF")
        if do_mc: selected_analyses.append("Monte Carlo")
        if do_mult: selected_analyses.append("Multiples")
        if do_scen: selected_analyses.append("Scenarios")
        if do_sens: selected_analyses.append("Sensitivity")
        
        st.session_state.analysis_type = selected_analyses
        
        # Reset button
        if st.button("Reset All", type="secondary"):
            st.session_state.valuation_results = None
            st.rerun()
    
    # Main content area
    if not selected_analyses:
        st.warning("Please select at least one analysis type from the sidebar.")
        return
    
    # Input section
    st.header("Input Parameters")
    
    # Create tabs for different input sections
    tab1, tab2, tab3, tab4 = st.tabs(["Financial Projections", "Valuation Assumptions", "Advanced Analysis", "Results"])
    
    with tab1:
        st.subheader("Financial Projections")
        
        # Input mode selection
        input_mode = st.radio(
            "Choose input method:",
            ["Driver-based (Revenue/Margins)", "Direct FCF Series"],
            help="Driver-based: Enter revenue and financial drivers\nDirect FCF: Enter free cash flows directly"
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            if input_mode == "Driver-based (Revenue/Margins)":
                st.write("**Revenue & Operating Metrics**")
                revenue_input = st.text_area(
                    "Revenue Series (comma-separated)",
                    value="100000000,110000000,120000000,130000000,140000000",
                    help="Enter projected revenues for each year, separated by commas. Use full numbers, e.g., 1000000 for one million."
                )
                
                ebit_margin = st.slider(
                    "EBIT Margin (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=20.0,
                    step=1.0,
                    help="Expected EBIT margin as percentage of revenue"
                ) / 100
                
                capex_input = st.text_area(
                    "Capital Expenditure Series",
                    value="10000000,11000000,12000000,13000000,14000000",
                    help="Enter projected CapEx for each year. Use full numbers."
                )
                
                depreciation_input = st.text_area(
                    "Depreciation Series",
                    value="5000000,6000000,7000000,8000000,9000000",
                    help="Enter projected depreciation for each year. Use full numbers."
                )
                
                nwc_input = st.text_area(
                    "Net Working Capital Changes",
                    value="2000000,2000000,2000000,2000000,2000000",
                    help="Enter projected NWC changes for each year. Use full numbers."
                )
                
                fcf_series = []
            else:
                st.write("**Direct FCF Input**")
                fcf_input = st.text_area(
                    "Free Cash Flow Series (comma-separated)",
                    value="50000000,55000000,60000000,65000000,70000000",
                    help="Enter projected free cash flows for each year. Use full numbers."
                )
                revenue_input = capex_input = depreciation_input = nwc_input = ""
                fcf_series = []
        
        with col2:
            st.write("**Capital Structure**")
            share_count = st.number_input(
                "Number of Shares Outstanding",
                min_value=1.0,
                value=100000000.0,
                step=1000000.0,
                help="Enter the total number of shares (e.g., 100000000 for 100 million)"
            )
            
            cost_of_debt = st.number_input(
                "Cost of Debt (%)",
                min_value=0.0,
                max_value=50.0,
                value=5.0,
                step=0.1,
                help="Cost of debt (leave at 0 to use WACC)"
            ) / 100
            # Debt schedule input (user-friendly)
            debt_schedule_input = create_user_friendly_debt_input(revenue_input)
    
    with tab2:
        st.subheader("Valuation Assumptions")
        
        col1, col2 = st.columns(2)
        
        with col1:
            wacc = st.slider(
                "WACC (%)",
                min_value=0.0,
                max_value=50.0,
                value=10.0,
                step=0.1,
                help="Weighted Average Cost of Capital"
            ) / 100
            
            tax_rate = st.slider(
                "Tax Rate (%)",
                min_value=0.0,
                max_value=100.0,
                value=21.0,
                step=1.0,
                help="Effective tax rate"
            ) / 100
            
        with col2:
            terminal_growth = st.slider(
                "Terminal Growth Rate (%)",
                min_value=0.0,
                max_value=10.0,
                value=2.0,
                step=0.1,
                help="Long-term growth rate for terminal value"
            ) / 100
            
            mid_year_convention = st.checkbox(
                "Mid-Year Convention",
                value=False,
                help="Check if cash flows occur mid-year (uncheck for year-end)"
            )
    
    with tab3:
        st.subheader("Advanced Analysis Parameters")
        
        # Show instructions if no advanced analyses are selected
        if not any([do_mc, do_mult, do_scen, do_sens]):
            st.info("Select advanced analyses from the sidebar to configure their parameters here.")
            st.write("""
            **Available Advanced Analyses:**
            - **Monte Carlo Simulation**: Uncertainty analysis with probability distributions
            - **Comparable Multiples**: Peer company analysis using industry ratios
            - **Scenario Analysis**: "What-if" testing with parameter overrides
            - **Sensitivity Analysis**: Parameter impact assessment
            """)
            
            # Initialize variables to avoid errors
            mc_runs = 2000
            mc_specs_input = '{}'
            comps_file = None
            scenarios_input = '{}'
            sensitivity_input = '{}'
        else:
            # Monte Carlo parameters
            if do_mc:
                st.write("**Monte Carlo Simulation**")
                mc_runs = st.number_input(
                    "Number of Simulations",
                    min_value=100,
                    max_value=10000,
                    value=2000,
                    step=100,
                    help="Number of Monte Carlo iterations"
                )
                
                mc_specs_input = st.text_area(
                    "Variable Specifications (JSON)",
                    value='{"wacc": {"dist": "normal", "params": {"loc": 0.10, "scale": 0.01}}}',
                    help="Define probability distributions for variables"
                )
            else:
                mc_runs = 2000
                mc_specs_input = '{}'
            
            # Multiples analysis
            if do_mult:
                st.write("**Comparable Companies**")
                comps_file = st.file_uploader(
                    "Upload Peer Comps CSV",
                    type="csv",
                    help="CSV file with peer company multiples"
                )
            else:
                comps_file = None
            
            # Scenarios
            if do_scen:
                st.write("**Scenario Analysis**")
                scenarios_input = json.dumps(create_user_friendly_scenario_input())
            else:
                scenarios_input = '{}'
            
            # Sensitivity analysis
            if do_sens:
                st.write("**Sensitivity Analysis**")
                sensitivity_input = json.dumps(create_user_friendly_sensitivity_input())
            else:
                sensitivity_input = '{}'
    
    # Initialize variables
    revenue = []
    capex = []
    depreciation = []
    nwc_changes = []
    fcf_series = []
    
    # Parse inputs and validate
    try:
        # Parse financial series
        if input_mode == "Driver-based (Revenue/Margins)":
            revenue = [float(x.strip()) for x in revenue_input.split(",") if x.strip()]
            capex = [float(x.strip()) for x in capex_input.split(",") if x.strip()]
            depreciation = [float(x.strip()) for x in depreciation_input.split(",") if x.strip()]
            nwc_changes = [float(x.strip()) for x in nwc_input.split(",") if x.strip()]
            fcf_series = []
        else:
            fcf_series = [float(x.strip()) for x in fcf_input.split(",") if x.strip()]
            revenue = []
            capex = []
            depreciation = []
            nwc_changes = []
        
        # Parse debt schedule
        debt_schedule = json.loads(debt_schedule_input) if debt_schedule_input else {}
        debt_schedule = {int(k): float(v) for k, v in debt_schedule.items()}
        
        # Parse advanced parameters
        mc_specs = json.loads(mc_specs_input) if do_mc and mc_specs_input else {}
        scenarios = json.loads(scenarios_input) if do_scen and scenarios_input else {}
        sensitivity = json.loads(sensitivity_input) if do_sens and sensitivity_input else {}
        
        # Check for circular references in scenarios
        if scenarios:
            circular_warnings = detect_circular_references(scenarios)
            for warning in circular_warnings:
                st.warning(warning)
        
        # Validate series lengths
        if input_mode == "Driver-based (Revenue/Margins)":
            series_lengths = [len(revenue), len(capex), len(depreciation), len(nwc_changes)]
            if len(set(series_lengths)) > 1:
                st.error("All financial series must have the same length")
                return
            
            # Additional validation for meaningful data
            if len(revenue) == 0:
                st.error("Revenue series cannot be empty")
                return
            if any(r <= 0 for r in revenue):
                st.error("Revenue values must be positive")
                return
                
        elif len(fcf_series) == 0:
            st.error("Please enter valid FCF series")
            return
        
    except (ValueError, json.JSONDecodeError) as e:
        st.error(f"Error parsing inputs: {str(e)}")
        return
    
    # Create ValuationParams object
    params = ValuationParams(
        revenue=revenue,
        ebit_margin=ebit_margin,  # This will now use the correct value from the UI slider
        capex=capex,
        depreciation=depreciation,
        nwc_changes=nwc_changes,
        fcf_series=fcf_series,
        terminal_growth=terminal_growth,
        wacc=wacc,
        tax_rate=tax_rate,
        mid_year_convention=mid_year_convention,
        share_count=share_count,
        cost_of_debt=cost_of_debt,
        debt_schedule=debt_schedule,
        variable_specs=mc_specs,
        scenarios=scenarios,
        sensitivity_ranges=sensitivity
    )
    
    # Terminal value sanity checks
    if terminal_growth > 0.05:  # 5% growth
        st.warning(
            f"⚠️ High terminal growth rate ({terminal_growth:.1%}). "
            "Consider whether this growth rate is sustainable in perpetuity."
        )
    
    if terminal_growth < -0.02:  # -2% growth
        st.warning(
            f"⚠️ Negative terminal growth rate ({terminal_growth:.1%}). "
            "This implies the business will shrink in perpetuity."
        )
    
    # Results tab
    with tab4:
        st.subheader("Valuation Results")
        
        if st.button("Run Valuation", type="primary"):
            with st.spinner("Running valuation analysis..."):
                results = run_valuation_analyses(params, selected_analyses, comps_file if do_mult else None, mc_runs if do_mc else 2000)
                st.session_state.valuation_results = results
                st.success("Valuation completed successfully!")
        
        # Display results
        if st.session_state.valuation_results:
            display_results(st.session_state.valuation_results, params)

def parse_series_input(input_text: str, name: str) -> list:
    """Parse comma-separated series input with error handling"""
    try:
        return [float(x.strip()) for x in input_text.split(",") if x.strip()]
    except ValueError:
        st.error(f"Invalid numbers in {name}. Please use comma-separated values.")
        return []

def create_user_friendly_debt_input(revenue_input):
    """Create a user-friendly debt schedule input interface"""
    st.write("**Debt Schedule (Optional)**")
    with st.expander("What is a debt schedule?", expanded=False):
        st.write("""
        **Debt Schedule** shows how much debt the company has each year.
        - **Year 0** = Current debt (today)
        - **Year 1** = Debt at end of year 1
        - **Year 2** = Debt at end of year 2
        - etc.
        
        **Example**: If you have $100M debt today that you plan to pay down:
        - Year 0: $100M
        - Year 1: $80M  
        - Year 2: $60M
        - Year 3: $40M
        """)
    debt_input_method = st.radio(
        "Choose debt input method:",
        ["Simple (Current debt only)", "Multi-year schedule"],
        help="Select how you want to enter debt information"
    )
    # Sample default schedule: 100.0, 80.0, 60.0, 40.0, 20.0, 0.0 (for 5 years)
    sample_defaults = [100000000.0, 80000000.0, 60000000.0, 40000000.0, 20000000.0, 0.0]
    if debt_input_method == "Simple (Current debt only)":
        current_debt = st.number_input(
            "Current Debt ($)",
            min_value=0.0,
            value=100000000.0,
            step=1000000.0,
            help="Enter the total debt in dollars, e.g., 1000000 for one million."
        )
        return json.dumps({0: current_debt})
    else:
        num_years = len([x for x in revenue_input.split(",") if x.strip()]) if revenue_input else 5
        debt_by_year = {}
        cols = st.columns(min(4, num_years + 1))
        for i in range(num_years + 1):
            with cols[i % len(cols)]:
                year_label = "Today" if i == 0 else f"Year {i}"
                default_val = sample_defaults[i] if i < len(sample_defaults) else 0.0
                debt_amount = st.number_input(
                    year_label,
                    min_value=0.0,
                    value=default_val,
                    step=1000000.0,
                    key=f"debt_{i}"
                )
                debt_by_year[i] = debt_amount
        if any(debt_by_year.values()):
            st.write("**Your Debt Schedule:**")
            debt_df = pd.DataFrame([
                {"Year": "Today" if k == 0 else f"Year {k}", "Debt ($)": v}
                for k, v in debt_by_year.items() if v > 0
            ])
            if not debt_df.empty:
                st.dataframe(debt_df, use_container_width=True)
        return json.dumps(debt_by_year)

def create_user_friendly_sensitivity_input():
    """Create a user-friendly sensitivity analysis input interface"""
    st.write("Sensitivity Analysis")
    with st.expander("What is sensitivity analysis?", expanded=False):
        st.write("""
        **Sensitivity Analysis** tests how changes in key parameters affect your valuation.
        **Example**: Test how WACC changes from 8% to 12% affect enterprise value.
        This helps identify which parameters have the biggest impact on your valuation.
        """)
    sensitivity_params = {}
    if st.checkbox("Test WACC sensitivity", value=True):
        st.write("**WACC Range (%)**")
        wacc_min = st.slider("Minimum WACC", 5.0, 15.0, 8.0, 0.5, key="wacc_min")
        wacc_max = st.slider("Maximum WACC", 5.0, 15.0, 12.0, 0.5, key="wacc_max")
        wacc_steps = st.slider("Number of steps", 3, 10, 5, key="wacc_steps")
        if wacc_min < wacc_max:
            wacc_range = [round(wacc_min + i * (wacc_max - wacc_min) / (wacc_steps - 1), 1) for i in range(wacc_steps)]
            sensitivity_params["wacc"] = [x/100 for x in wacc_range]
            st.write(f"**WACC values to test:** {wacc_range}%")
    if st.checkbox("Test Terminal Growth sensitivity", value=False):
        st.write("**Terminal Growth Range (%)**")
        growth_min = st.slider("Minimum Growth", -2.0, 5.0, 1.0, 0.5, key="growth_min")
        growth_max = st.slider("Maximum Growth", -2.0, 5.0, 3.0, 0.5, key="growth_max")
        growth_steps = st.slider("Number of steps", 3, 8, 5, key="growth_steps")
        if growth_min < growth_max:
            growth_range = [round(growth_min + i * (growth_max - growth_min) / (growth_steps - 1), 1) for i in range(growth_steps)]
            sensitivity_params["terminal_growth"] = [x/100 for x in growth_range]
            st.write(f"**Growth values to test:** {growth_range}%")
    if st.checkbox("Test EBIT Margin sensitivity", value=False):
        st.write("**EBIT Margin Range (%)**")
        margin_min = st.slider("Minimum Margin", 10.0, 30.0, 15.0, 1.0, key="margin_min")
        margin_max = st.slider("Maximum Margin", 10.0, 30.0, 25.0, 1.0, key="margin_max")
        margin_steps = st.slider("Number of steps", 3, 8, 5, key="margin_steps")
        if margin_min < margin_max:
            margin_range = [round(margin_min + i * (margin_max - margin_min) / (margin_steps - 1), 1) for i in range(margin_steps)]
            sensitivity_params["ebit_margin"] = [x/100 for x in margin_range]
            st.write(f"**Margin values to test:** {margin_range}%")
    return sensitivity_params

def create_user_friendly_scenario_input():
    """Create a user-friendly scenario analysis input interface"""
    st.write("Scenario Analysis")
    with st.expander("What is scenario analysis?", expanded=False):
        st.write("""
        **Scenario Analysis** tests different "what-if" situations.
        **Example Scenarios:**
        - **Base Case**: Your current assumptions
        - **Optimistic**: Better performance (higher margins, growth)
        - **Pessimistic**: Worse performance (lower margins, growth)
        """)
    scenarios = {}
    scenarios["Base"] = {}
    if st.checkbox("Add Optimistic scenario", value=True):
        st.write("**Optimistic Scenario Parameters**")
        opt_ebit_margin = st.slider("EBIT Margin (%)", 15.0, 35.0, 25.0, 1.0, key="opt_margin")
        opt_growth = st.slider("Terminal Growth (%)", 1.0, 5.0, 3.0, 0.5, key="opt_growth")
        opt_wacc = st.slider("WACC (%)", 8.0, 12.0, 9.0, 0.5, key="opt_wacc")
        scenarios["Optimistic"] = {
            "ebit_margin": opt_ebit_margin / 100,
            "terminal_growth": opt_growth / 100,
            "wacc": opt_wacc / 100
        }
    if st.checkbox("Add Pessimistic scenario", value=True):
        st.write("**Pessimistic Scenario Parameters**")
        pes_ebit_margin = st.slider("EBIT Margin (%)", 10.0, 25.0, 15.0, 1.0, key="pes_margin")
        pes_growth = st.slider("Terminal Growth (%)", -1.0, 3.0, 1.0, 0.5, key="pes_growth")
        pes_wacc = st.slider("WACC (%)", 10.0, 15.0, 12.0, 0.5, key="pes_wacc")
        scenarios["Pessimistic"] = {
            "ebit_margin": pes_ebit_margin / 100,
            "terminal_growth": pes_growth / 100,
            "wacc": pes_wacc / 100
        }
    return scenarios


def run_valuation_analyses(params: ValuationParams, analyses: list, comps_file=None, mc_runs=2000) -> dict:
    """Run all selected valuation analyses"""
    results = {}
    
    # DCF Analysis
    if "WACC DCF" in analyses:
        try:
            ev, equity, ps = calc_dcf_series(params)
            results["wacc_dcf"] = {"EV": ev, "Equity": equity, "PS": ps}
        except Exception as e:
            st.error(f"WACC DCF calculation failed: {str(e)}")
    
    if "APV DCF" in analyses:
        try:
            ev, equity, ps = calc_apv(params)
            results["apv_dcf"] = {"EV": ev, "Equity": equity, "PS": ps}
        except Exception as e:
            st.error(f"APV calculation failed: {str(e)}")
    
    # Monte Carlo
    if "Monte Carlo" in analyses:
        try:
            # Use a fixed seed for reproducibility
            mc_results = run_monte_carlo(params, runs=mc_runs, random_seed=42)
            results["monte_carlo"] = mc_results
        except Exception as e:
            st.error(f"Monte Carlo simulation failed: {str(e)}")
    
    # Multiples Analysis
    if "Multiples" in analyses and comps_file:
        try:
            comps_df = pd.read_csv(comps_file)
            mult_results = run_multiples_analysis(params, comps_df)
            results["multiples"] = mult_results
        except Exception as e:
            st.error(f"Multiples analysis failed: {str(e)}")
    
    # Scenarios
    if "Scenarios" in analyses:
        try:
            scen_results = run_scenarios(params)
            results["scenarios"] = scen_results
        except Exception as e:
            st.error(f"Scenario analysis failed: {str(e)}")
    
    # Sensitivity
    if "Sensitivity" in analyses:
        try:
            sens_results = run_sensitivity_analysis(params)
            results["sensitivity"] = sens_results
        except Exception as e:
            st.error(f"Sensitivity analysis failed: {str(e)}")
    
    return results

def display_results(results: dict, params: ValuationParams):
    """Display valuation results with professional formatting"""
    
    # Summary metrics
    st.subheader("Summary Valuation")
    st.info("All values are in dollars (except price per share)")
    
    # Create summary table
    summary_data = []
    if "wacc_dcf" in results:
        summary_data.append({
            "Method": "WACC DCF",
            "Enterprise Value ($)": f"${results['wacc_dcf']['EV']:,.0f}",
            "Equity Value ($)": f"${results['wacc_dcf']['Equity']:,.0f}",
            "Price per Share ($)": f"${results['wacc_dcf']['PS']:,.2f}" if results['wacc_dcf']['PS'] else "N/A"
        })
    
    if "apv_dcf" in results:
        summary_data.append({
            "Method": "APV DCF",
            "Enterprise Value ($)": f"${results['apv_dcf']['EV']:,.0f}",
            "Equity Value ($)": f"${results['apv_dcf']['Equity']:,.0f}",
            "Price per Share ($)": f"${results['apv_dcf']['PS']:,.2f}" if results['apv_dcf']['PS'] else "N/A"
        })
    
    if summary_data:
        summary_df = pd.DataFrame(summary_data)
        st.dataframe(summary_df, use_container_width=True)
    
    # Detailed results in expandable sections
    if "wacc_dcf" in results or "apv_dcf" in results:
        with st.expander("DCF Analysis Details", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                if "wacc_dcf" in results:
                    st.metric("WACC DCF - Enterprise Value", f"${results['wacc_dcf']['EV']:,.0f}")
                    st.metric("WACC DCF - Equity Value", f"${results['wacc_dcf']['Equity']:,.0f}")
                    if results['wacc_dcf']['PS']:
                        st.metric("WACC DCF - Price per Share", f"${results['wacc_dcf']['PS']:,.2f}")
            
            with col2:
                if "apv_dcf" in results:
                    st.metric("APV - Enterprise Value", f"${results['apv_dcf']['EV']:,.0f}")
                    st.metric("APV - Equity Value", f"${results['apv_dcf']['Equity']:,.0f}")
                    if results['apv_dcf']['PS']:
                        st.metric("APV - Price per Share", f"${results['apv_dcf']['PS']:,.2f}")
    
    # Monte Carlo Results
    if "monte_carlo" in results:
        with st.expander("Monte Carlo Simulation", expanded=True):
            mc_data = results["monte_carlo"]
            
            if "WACC" in mc_data:
                wacc_df = mc_data["WACC"]
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**WACC DCF Distribution Statistics**")
                    stats_df = wacc_df.describe()
                    st.dataframe(stats_df)
                
                with col2:
                    # Create histogram
                    if PLOTLY_AVAILABLE:
                        fig = px.histogram(
                            wacc_df, 
                            x="EV", 
                            nbins=30,
                            title="Enterprise Value Distribution (WACC DCF)",
                            labels={"EV": "Enterprise Value ($)", "count": "Frequency"}
                        )
                        fig.update_layout(showlegend=False)
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        # Fallback to matplotlib
                        fig, ax = plt.subplots()
                        ax.hist(wacc_df["EV"], bins=30)
                        ax.set_title("Enterprise Value Distribution (WACC DCF)")
                        ax.set_xlabel("Enterprise Value ($)")
                        ax.set_ylabel("Frequency")
                        st.pyplot(fig)
                
                # Confidence intervals
                st.write("**Confidence Intervals**")
                percentiles = [5, 25, 50, 75, 95]
                ci_data = []
                for p in percentiles:
                    ci_data.append({
                        "Percentile": f"{p}%",
                        "Enterprise Value ($)": f"${wacc_df['EV'].quantile(p/100):,.1f}",
                        "Equity Value ($)": f"${wacc_df['Equity'].quantile(p/100):,.1f}",
                        "Price per Share ($)": f"${wacc_df['PS'].quantile(p/100):,.2f}" if 'PS' in wacc_df.columns else "N/A"
                    })
                
                ci_df = pd.DataFrame(ci_data)
                st.dataframe(ci_df, use_container_width=True)
    
    # Multiples Analysis
    if "multiples" in results:
        with st.expander("Comparable Multiples Analysis", expanded=True):
            mult_df = results["multiples"]
            st.dataframe(mult_df, use_container_width=True)
            
            # Summary statistics
            if not mult_df.empty:
                mean_ev = mult_df["Mean Implied EV"].mean()
                median_ev = mult_df["Median Implied EV"].median()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Average Implied EV", f"${mean_ev:,.0f}")
                with col2:
                    st.metric("Median Implied EV", f"${median_ev:,.0f}")
    
    # Scenario Analysis
    if "scenarios" in results:
        with st.expander("Scenario Analysis", expanded=True):
            scen_df = results["scenarios"]
            st.dataframe(scen_df, use_container_width=True)
            
            # Scenario comparison chart
            if not scen_df.empty:
                if PLOTLY_AVAILABLE:
                    fig = go.Figure()
                    fig.add_trace(go.Bar(
                        x=scen_df.index,
                        y=scen_df["EV"],
                        name="Enterprise Value",
                        marker_color='lightblue'
                    ))
                    fig.update_layout(
                        title="Enterprise Value by Scenario",
                        xaxis_title="Scenario",
                        yaxis_title="Enterprise Value ($)",
                        showlegend=False
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    # Fallback to matplotlib
                    fig, ax = plt.subplots()
                    ax.bar(scen_df.index, scen_df["EV"], color='lightblue')
                    ax.set_title("Enterprise Value by Scenario")
                    ax.set_xlabel("Scenario")
                    ax.set_ylabel("Enterprise Value ($)")
                    plt.xticks(rotation=45)
                    st.pyplot(fig)
    
    # Sensitivity Analysis
    if "sensitivity" in results:
        with st.expander("Sensitivity Analysis", expanded=True):
            sens_df = results["sensitivity"]
            st.dataframe(sens_df, use_container_width=True)
            
            # Sensitivity heatmap
            if not sens_df.empty:
                if PLOTLY_AVAILABLE:
                    fig = px.imshow(
                        sens_df.T,
                        title="Sensitivity Heatmap",
                        labels=dict(x="Parameter Values", y="Parameters", color="Enterprise Value ($)"),
                        aspect="auto"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    # Fallback to matplotlib
                    fig, ax = plt.subplots()
                    im = ax.imshow(sens_df.T, aspect='auto', cmap='viridis')
                    ax.set_title("Sensitivity Heatmap")
                    ax.set_xlabel("Parameter Values")
                    ax.set_ylabel("Parameters")
                    plt.colorbar(im, ax=ax, label="Enterprise Value ($)")
                    st.pyplot(fig)
    
    # Download functionality
    st.subheader("Download Results")
    
    # Create downloadable data
    download_data = {}
    
    # Summary data
    if summary_data:
        download_data["summary"] = pd.DataFrame(summary_data)
    
    # Detailed results
    if "monte_carlo" in results:
        for method, df in results["monte_carlo"].items():
            download_data[f"monte_carlo_{method.lower()}"] = df
    
    if "multiples" in results:
        download_data["multiples"] = results["multiples"]
    
    if "scenarios" in results:
        download_data["scenarios"] = results["scenarios"]
    
    if "sensitivity" in results:
        download_data["sensitivity"] = results["sensitivity"]
    
    # Create Excel file
    if OPENPYXL_AVAILABLE and download_data:
        # Create a temporary file-like object
        buffer = io.BytesIO()
        
        # Write to Excel using openpyxl directly
        wb = Workbook()
        
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        # Add sheets for each dataset
        for sheet_name, df in download_data.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
            
            # Write headers
            for col, header in enumerate(df.columns, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data efficiently using batch operations
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to buffer and clean up
        wb.save(buffer)
        buffer.seek(0)
        wb.close()  # Explicitly close to free memory
        
        st.download_button(
            label="Download Excel Report",
            data=buffer.getvalue(),
            file_name=f"valuation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    elif download_data and not OPENPYXL_AVAILABLE:
        st.warning("Excel export not available. Use CSV download instead.")
    
    # Create CSV summary
    if summary_data:
        csv_data = pd.DataFrame(summary_data).to_csv(index=False)
        st.download_button(
            label="Download Summary CSV",
            data=csv_data,
            file_name=f"valuation_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )

if __name__ == "__main__":
    main()
